
import os
from openpyxl import load_workbook

dataset_root = '/mnt/ourDataset_v2/ourDataset_v2_label'
scenes = os.listdir(dataset_root)
scenes.sort()

xlsx_path = os.path.join('/mnt/ChillDisk/personal_data/zhangq/RadarRGBFusionNet2/GroupPath', '20230919AllData.xlsx')
wb = load_workbook(xlsx_path)  # type:workbook
ws = wb.active
num = 0
for scene in scenes:
    num += 1
    # print('scene', scene)
    # group_name = [scene, num]

    for j, c in enumerate("A"):
        ws[f"{c}{num}"] = str(scene)

wb.save(xlsx_path)


